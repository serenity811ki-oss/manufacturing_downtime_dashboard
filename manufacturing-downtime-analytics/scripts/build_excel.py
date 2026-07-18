import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

SRC = "/home/claude/project/data/manufacturing_downtime_data.csv"
OUT = "/home/claude/project/data/Manufacturing_Downtime_Dataset.xlsx"

df = pd.read_csv(SRC)

# ---------------------------------------------------------------------
# Build a Date dimension table (Jan 2023 - Dec 2024) for the Power BI model
# ---------------------------------------------------------------------
start = datetime(2023, 1, 1)
end = datetime(2024, 12, 31)
dates = pd.date_range(start, end, freq="D")
date_dim = pd.DataFrame({"Date": dates})
date_dim["Year"] = date_dim["Date"].dt.year
date_dim["Month Number"] = date_dim["Date"].dt.month
date_dim["Month"] = date_dim["Date"].dt.strftime("%B")
date_dim["Month Short"] = date_dim["Date"].dt.strftime("%b")
date_dim["Year-Month"] = date_dim["Date"].dt.strftime("%Y-%m")
date_dim["Week"] = date_dim["Date"].dt.isocalendar().week.astype(int)
date_dim["Day Name"] = date_dim["Date"].dt.strftime("%A")
date_dim["Day Number"] = date_dim["Date"].dt.day
date_dim["Quarter"] = "Q" + date_dim["Date"].dt.quarter.astype(str)
date_dim["Is Weekend"] = date_dim["Date"].dt.dayofweek.isin([5, 6])
date_dim["Date"] = date_dim["Date"].dt.strftime("%Y-%m-%d")

# ---------------------------------------------------------------------
# Machine dimension (deduped) - supports a clean star-schema in Power BI
# ---------------------------------------------------------------------
machine_dim = (
    df[["Machine ID", "Machine Type", "Plant", "Production Line"]]
    .drop_duplicates()
    .sort_values(["Plant", "Production Line", "Machine ID"])
    .reset_index(drop=True)
)

# ---------------------------------------------------------------------
# Write workbook
# ---------------------------------------------------------------------
wb = Workbook()

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Arial", size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def write_sheet(ws, dataframe):
    for c_idx, col in enumerate(dataframe.columns, start=1):
        cell = ws.cell(row=1, column=c_idx, value=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
    for r_idx, row in enumerate(dataframe.itertuples(index=False), start=2):
        for c_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            cell.font = BODY_FONT
            cell.border = BORDER
    # Column widths
    for c_idx, col in enumerate(dataframe.columns, start=1):
        if len(dataframe):
            max_len = max(len(str(v)) for v in dataframe[col].tolist())
        else:
            max_len = 0
        max_len = max(max_len, len(str(col)))
        ws.column_dimensions[get_column_letter(c_idx)].width = min(max(max_len + 3, 10), 32)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

# Sheet 1: Fact table
ws_fact = wb.active
ws_fact.title = "Fact_Downtime_Events"
write_sheet(ws_fact, df)

# Sheet 2: Date dimension
ws_date = wb.create_sheet("Dim_Date")
write_sheet(ws_date, date_dim)

# Sheet 3: Machine dimension
ws_machine = wb.create_sheet("Dim_Machine")
write_sheet(ws_machine, machine_dim)

# Sheet 4: Data dictionary
dict_rows = [
    ("Event ID", "Unique identifier for each downtime event", "Text", "Fact_Downtime_Events"),
    ("Date", "Calendar date the downtime event occurred", "Date", "Fact_Downtime_Events (FK to Dim_Date)"),
    ("Year", "Calendar year", "Whole Number", "Fact_Downtime_Events"),
    ("Month", "Month name", "Text", "Fact_Downtime_Events"),
    ("Month Number", "Numeric month (1-12), used for chronological sorting", "Whole Number", "Fact_Downtime_Events"),
    ("Week", "ISO week number", "Whole Number", "Fact_Downtime_Events"),
    ("Plant", "Manufacturing plant/site", "Text", "Fact_Downtime_Events"),
    ("Production Line", "Line within the plant", "Text", "Fact_Downtime_Events"),
    ("Machine ID", "Unique machine/asset identifier", "Text", "Fact_Downtime_Events (FK to Dim_Machine)"),
    ("Machine Type", "Category of equipment", "Text", "Fact_Downtime_Events"),
    ("Shift", "Working shift when the event occurred", "Text", "Fact_Downtime_Events"),
    ("Operator ID", "Operator on duty", "Text", "Fact_Downtime_Events"),
    ("Technician", "Maintenance technician who handled the event", "Text", "Fact_Downtime_Events"),
    ("Downtime Category", "High-level downtime classification", "Text", "Fact_Downtime_Events"),
    ("Downtime Reason", "Specific reason within the category", "Text", "Fact_Downtime_Events"),
    ("Planned or Unplanned", "Whether the downtime was scheduled", "Text", "Fact_Downtime_Events"),
    ("Root Cause", "Root cause narrative from the maintenance log", "Text", "Fact_Downtime_Events"),
    ("Downtime Minutes", "Duration of the downtime event in minutes", "Decimal", "Fact_Downtime_Events"),
    ("Downtime Hours", "Duration of the downtime event in hours (Minutes / 60)", "Decimal", "Fact_Downtime_Events"),
    ("Repair Cost EURO", "Total cost of parts + labor for the repair, in EUR", "Decimal (Currency)", "Fact_Downtime_Events"),
    ("Cost per Minute", "Repair Cost EURO / Downtime Minutes", "Decimal (Currency)", "Fact_Downtime_Events"),
    ("Production Units Lost", "Estimated production units lost due to the event", "Whole Number", "Fact_Downtime_Events"),
    ("Product Type", "Product family running on the line at the time", "Text", "Fact_Downtime_Events"),
    ("Status", "Resolved / In Progress / Pending", "Text", "Fact_Downtime_Events"),
    ("Resolution Time", "Hours between event start and full resolution (blank if Pending)", "Decimal", "Fact_Downtime_Events"),
    ("Maintenance Type", "Corrective / Preventive / Predictive / Emergency / N/A", "Text", "Fact_Downtime_Events"),
]
dict_df = pd.DataFrame(dict_rows, columns=["Column Name", "Description", "Data Type", "Location"])
ws_dict = wb.create_sheet("Data_Dictionary")
write_sheet(ws_dict, dict_df)

# Reorder: Fact first, then dims, then dictionary
wb._sheets = [ws_fact, ws_date, ws_machine, ws_dict]

wb.save(OUT)
print("Saved workbook:", OUT)
print("Fact rows:", len(df), " | Date dim rows:", len(date_dim), " | Machine dim rows:", len(machine_dim))
